import openpyxl

from .models import MeasurePoint, Customer, MeasuresSetting
from datetime import date
import os


class ExportData:
    template_filename = 'vitrek/protocol_template.xlsm'
    ps_name = 'Протокол'  # protocol title sheet name
    ds_sheet_name = 'Измерения'  # measures data sheet name

    dev_type_addr = 'C4'
    dev_num_addr = 'C6'

    start_column = 2
    point_row = 2
    date_row = 7
    curr_type_row = 6
    meas_start_row = 13

    @classmethod
    def export_data(cls):
        customer = Customer.get_customer()
        points = MeasurePoint.get_measures_data()

        wb = openpyxl.open(cls.template_filename, keep_vba=True, read_only=False)
        ws = wb[cls.ps_name]
        ws[cls.dev_type_addr] = customer.type
        ws[cls.dev_num_addr] = customer.number

        ws = wb[cls.ds_sheet_name]

        for col, point in enumerate(points, start=cls.start_column):
            ws.cell(cls.point_row, col).value = point.measure_point
            ws.cell(cls.curr_type_row, col).value = point.current_type  # change selection AC, DC, AC_0,1

             # insert measures
            for row, data in enumerate(point.measures['data'], start=cls.meas_start_row):
                if point.current_type == 'переменный':
                    ws.cell(row, col).value = data['ACV']
                elif point.current_type == 'постоянный':
                    ws.cell(row, col).value = data['DCV']
                elif point.current_type == 'переменный 0,1 Гц':
                    ws.cell(row, col).value = data['ACV']
        wb.save(cls.get_file_address())

    @classmethod
    def gen_file_name(cls):
        customer = Customer.get_customer()
        return f'{customer.name} {customer.type} №{customer.number}.xlsm'

    @classmethod
    def get_file_address(cls):
        path = MeasuresSetting.get_file_path()
        f_name = cls.gen_file_name()
        return os.path.join(path, f_name)


